"""One-off script to apply CyberShield schema fixes. Run: python scripts/update_api_schema.py"""

from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl

SCHEMA_PATH = Path(__file__).resolve().parents[1] / "docs" / "CyberShield_NIC_API_Schema.xlsx"


def find_row(ws, col: int, text: str, start: int = 1) -> int | None:
    for r in range(start, ws.max_row + 1):
        val = ws.cell(r, col).value
        if val and text in str(val):
            return r
    return None


def set_cell(ws, row: int, col: int, value) -> None:
    ws.cell(row, col, value)


def main() -> None:
    wb = openpyxl.load_workbook(SCHEMA_PATH)

    # --- Conventions: point error.code to Error Responses sheet ---
    conv = wb["Conventions"]
    for r in range(1, conv.max_row + 1):
        if conv.cell(r, 1).value == "Rule" and conv.cell(r, 2).value:
            text = str(conv.cell(r, 2).value)
            if "see HTTP status sheet" in text:
                set_cell(
                    conv,
                    r,
                    2,
                    "error.code is a fixed uppercase string (see Error Responses sheet), "
                    "error.message is human-readable, safe to show in UI",
                )

    # --- REST Endpoints: attribution narrative + audit trail actor fields ---
    rest = wb["REST Endpoints"]
    for r in range(1, rest.max_row + 1):
        endpoint = str(rest.cell(r, 3).value or "")
        resp = str(rest.cell(r, 6).value or "")

        if "/attributions/{attribution_id}" in endpoint:
            if "narrative" not in resp:
                new_resp = resp.replace(
                    "recommended_actions: [ string ]",
                    "narrative: string, recommended_actions: [ string ]",
                )
                if new_resp == resp:
                    new_resp = (
                        "{ attribution_id, anomaly_id, mitre_technique_id, mitre_tactic, "
                        "matched_campaign, confidence, narrative, "
                        "predicted_next_techniques: [ { technique_id, tactic, likelihood } ], "
                        "recommended_actions: [ string ] }"
                    )
                set_cell(rest, r, 6, new_resp)

        if endpoint == "/audit-trail":
            if "actor_id" not in resp:
                new_resp = (
                    "{ items: [ { audit_id, agent, actor_id, actor_name, action_summary, "
                    "reasoning, confidence, timestamp } ], total, limit, offset }"
                )
                set_cell(rest, r, 6, new_resp)
            req = str(rest.cell(r, 5).value or "")
            if "*limit" in req:
                set_cell(rest, r, 5, "query: limit, offset, agent, from, to (all optional)")

        if endpoint == "/dashboard/summary":
            if "threat_posture" in resp and "enum" not in resp.lower():
                new_resp = resp.replace(
                    "threat_posture",
                    "threat_posture (enum: see Enums sheet)",
                )
                set_cell(rest, r, 6, new_resp)

    # --- REST Endpoints: fix *limit/*offset markers on list endpoints ---
    for r in range(1, rest.max_row + 1):
        req = str(rest.cell(r, 5).value or "")
        if "*limit" in req or "*offset" in req:
            set_cell(
                rest,
                r,
                5,
                req.replace("*limit", "limit").replace("*offset", "offset")
                + " (limit/offset optional; defaults apply)",
            )
    pag = wb["Pagination"]
    for r in range(1, pag.max_row + 1):
        label = str(pag.cell(r, 1).value or "")
        if label in ("limit", "offset"):
            set_cell(pag, r, 3, "No")
            if label == "limit":
                set_cell(
                    pag,
                    r,
                    4,
                    "Optional. Default 20 if omitted. Max 100. Server applies default — no 400.",
                )
            else:
                set_cell(
                    pag,
                    r,
                    4,
                    "Optional. Default 0 if omitted. Zero-indexed. Server applies default — no 400.",
                )

    # Add clarification row if not present
    if not any(
        "limit and offset are OPTIONAL" in str(pag.cell(r, 2).value or "")
        for r in range(1, pag.max_row + 1)
    ):
        next_row = pag.max_row + 1
        set_cell(pag, next_row, 1, "Note")
        set_cell(
            pag,
            next_row,
            2,
            "limit and offset are OPTIONAL. If omitted, server uses defaults (limit=20, offset=0). "
            "GET /anomalies with no query string returns 200 with first page.",
        )

    # --- Query Parameters: pagination optional; agent enum note ---
    qp = wb["Query Parameters"]
    for r in range(1, qp.max_row + 1):
        param = str(qp.cell(r, 2).value or "")
        if param in ("limit", "offset"):
            set_cell(qp, r, 4, "No")
        if param == "agent":
            set_cell(
                qp,
                r,
                5,
                "Filter: analyst | ml_engine | orchestrator (see Enums sheet). "
                "Note: 'human' deprecated — use analyst for human SOC actions.",
            )

    # --- Enums sheet: add missing enums ---
    enums = wb["Enums & Status Codes"]
    existing = {
        str(enums.cell(r, 1).value or "").strip().lower()
        for r in range(1, enums.max_row + 1)
    }

    new_enums = [
        ("threat_posture", "nominal, elevated, high, critical", "/dashboard/summary"),
        (
            "audit_agent",
            "analyst, ml_engine, orchestrator",
            "audit-trail (agent field). analyst = human SOC analyst action",
        ),
    ]

    row = enums.max_row + 1
    for enum_name, values, used_on in new_enums:
        if enum_name not in existing:
            set_cell(enums, row, 1, enum_name)
            set_cell(enums, row, 2, values)
            set_cell(enums, row, 3, used_on)
            row += 1

    # Update HTTP status codes pointer in Enums sheet
    for r in range(1, enums.max_row + 1):
        if enums.cell(r, 1).value == "200" and enums.cell(r, 2).value == "OK":
            set_cell(
                enums,
                r - 1,
                1,
                "HTTP status codes (see also Error Responses sheet for error.code strings)",
            )
            break

    # --- Best Practices: caching guidance ---
    bp = wb["Best Practices"]
    notes = [
        "",
        "11. Caching guidance (live SOC dashboard)",
        "SAFE to cache (low churn): GET /assets, GET /cve-feed, GET /attributions/{id} (short TTL)",
        "DO NOT cache (live state): GET /anomalies, GET /actions, GET /dashboard/summary, WebSocket feeds",
        "Stale pending actions or anomaly counts are dangerous — use Cache-Control: no-store on live endpoints.",
    ]
    for note in notes:
        if note and any(note.split(":")[0] in str(bp.cell(r, 1).value or "") for r in range(1, bp.max_row + 1)):
            continue
        next_row = bp.max_row + 1
        set_cell(bp, next_row, 1, note)

    # --- Response Headers: note which endpoints get cache headers ---
    if "Response Headers" in wb.sheetnames:
        rh = wb["Response Headers"]
        if not any("Do not cache" in str(rh.cell(r, 1).value or "") for r in range(1, rh.max_row + 1)):
            start = rh.max_row + 2
            rows = [
                ("Caching policy", "", "", ""),
                (
                    "Cacheable GETs",
                    "Cache-Control",
                    "public, max-age=300",
                    "/assets, /cve-feed only",
                ),
                (
                    "Live GETs",
                    "Cache-Control",
                    "no-store",
                    "/anomalies, /actions, /dashboard/summary — never serve stale security state",
                ),
            ]
            for i, row_data in enumerate(rows, start=start):
                for c, val in enumerate(row_data, start=1):
                    set_cell(rh, i, c, val)

    wb.save(SCHEMA_PATH)
    print(f"Updated {SCHEMA_PATH}")


if __name__ == "__main__":
    main()
